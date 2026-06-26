import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
excel_file = "data/Engine Programme Management.xlsx"
event = "TC2K-2026-04"

LIMITS = {
    "Presión aceite mín FL": {
        "warning": 3.0,
        "critical": 2.5,
        "direction": "low",
    },
    "Temp. aceite máx FL": {
        "warning": 140,
        "critical": 150,
        "direction": "high",
    },
    "Temp. agua máx FL": {
        "warning": 105,
        "critical": 110,
        "direction": "high",
    },
    "Temp. máxima detectada agua": {
        "warning": 110,
        "critical": 115,
        "direction": "high",
    },
    "RPM turbo máx FL": {
        "warning": 165000,
        "critical": 175000,
        "direction": "high",
    },
    "Presión combustible mín FL": {
        "warning": 3.0,
        "critical": 2.5,
        "direction": "low",
    },
    "Temp. escape máx": {
        "warning": 990,
        "critical": 1100,
        "direction": "high",
    },
    "Voltaje batería mín FL": {
        "warning": 12.5,
        "critical": 11.5,
        "direction": "low",
    },
    
    "P Cárter máx FL": {
    "warning": 0,
    "critical": 20,
    "direction": "high",
},
}

def generar_observaciones(row):
    observaciones = []
    prioridad = "🟢 OK"

    for columna, reglas in LIMITS.items():
        valor = row[columna]

        if pd.isna(valor):
            continue

        warning = reglas["warning"]
        critical = reglas["critical"]
        direction = reglas["direction"]

        if direction == "high":
            if valor >= critical:
                observaciones.append(f"🔴 {columna} fuera de rango ({valor})")
                prioridad = "🔴 CRÍTICO"
            elif valor >= warning:
                observaciones.append(f"🟡 {columna} cerca del límite ({valor})")
                if prioridad != "🔴 CRÍTICO":
                    prioridad = "🟡 ATENCIÓN"

        elif direction == "low":
            if valor <= critical:
                observaciones.append(f"🔴 {columna} fuera de rango ({valor})")
                prioridad = "🔴 CRÍTICO"
            elif valor <= warning:
                observaciones.append(f"🟡 {columna} cerca del límite ({valor})")
                if prioridad != "🔴 CRÍTICO":
                    prioridad = "🟡 ATENCIÓN"

    if not observaciones:
        observaciones.append("✅ Sin observaciones")

    return prioridad, " | ".join(observaciones)

df = pd.read_excel(excel_file, sheet_name="radb")

df = df[df["Carrera"] == event]
columns = [
    "Carrera",
    "Run name",
    "Motor",
    "Piloto",
    "Auto",
    "pOil_FL_min",
    "tOil_FL_max",
    "tWater_FL_max",
    "tWater_max",
    "nTurbo_FL_max",
    "pFuel_FL_min",
    "tAir_FL_max",
    "tExhaust_max",
    "VBatt_FL_min",
    "pBoost_error_FL_max",
    "pCarter_FL_max",
]

report = df[columns]

print(report.head(20))

output_file = "reports/Weekend Analyzer.xlsx"

report = report.rename(
    columns={
        "Carrera": "Evento",
        "Run name": "Run",
        "pOil_FL_min": "Presión aceite mín FL",
        "tOil_FL_max": "Temp. aceite máx FL",
        "tWater_FL_max": "Temp. agua máx FL",
        "tWater_max": "Temp. máxima detectada agua",
        "nTurbo_FL_max": "RPM turbo máx FL",
        "pFuel_FL_min": "Presión combustible mín FL",
        "tAir_FL_max": "Temp. aire máx FL",
        "tExhaust_max": "Temp. escape máx",
        "VBatt_FL_min": "Voltaje batería mín FL",
        "pBoost_error_FL_max": "Error Boost máx FL",
        "pCarter_FL_max": "P Cárter máx FL",
    }
)
for column in LIMITS.keys():
    report[column] = pd.to_numeric(report[column], errors="coerce")

report[["Prioridad", "Observaciones"]] = report.apply(
    generar_observaciones,
    axis=1,
    result_type="expand"
)

def prioridad_num(valor):
    if "CRÍTICO" in valor:
        return 1
    elif "ATENCIÓN" in valor:
        return 2
    else:
        return 3
first_columns = [
    "Evento",
    "Run",
    "Motor",
    "Piloto",
    "Auto",
    "Prioridad",
    "Observaciones",
]

other_columns = [col for col in report.columns if col not in first_columns]

report = report[first_columns + other_columns]
report["prioridad_num"] = report["Prioridad"].apply(prioridad_num)

report = report.sort_values(
    by=["prioridad_num", "Run", "Motor"],
    ascending=[True, True, True]
)

report = report.drop(columns=["prioridad_num"])

alerts = report[
    report["Prioridad"] != "🟢 OK"
].copy()

alerts = report[
    report["Prioridad"] != "🟢 OK"
].copy()

if alerts.empty:
    alerts = pd.DataFrame({
        "Estado": [
            "✅ No se detectaron alertas para este evento."
        ]
    })

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    report.to_excel(
        writer,
        sheet_name="Weekend Analyzer",
        index=False
    )

    alerts.to_excel(
        writer,
        sheet_name="Resumen de Alertas",
        index=False
    )


wb = load_workbook(output_file)
ws = wb.active

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for cell in ws[1]:
    cell.font = Font(bold=True)

prioridad_column = None

for cell in ws[1]:
    if cell.value == "Prioridad":
        prioridad_column = cell.column
        break

for row in range(2, ws.max_row + 1):
    prioridad_cell = ws.cell(row=row, column=prioridad_column)

    if "OK" in str(prioridad_cell.value):
        prioridad_cell.fill = green_fill
    elif "ATENCIÓN" in str(prioridad_cell.value):
        prioridad_cell.fill = yellow_fill
    elif "CRÍTICO" in str(prioridad_cell.value):
        prioridad_cell.fill = red_fill

for column_cells in ws.columns:
    max_length = max(
        len(str(cell.value)) if cell.value is not None else 0
        for cell in column_cells
    )
    column_letter = get_column_letter(column_cells[0].column)
    ws.column_dimensions[column_letter].width = max_length + 2

wb.save(output_file)


print("✅ Weekend Analyzer generado correctamente.")

