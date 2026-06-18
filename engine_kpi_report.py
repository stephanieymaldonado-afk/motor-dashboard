from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

SERVICE_LIMIT = 2500
WARNING_LIMIT = 2250

excel_file = "data/Engine Programme Management.xlsx"
output_file = "reports/Motor Dashboard.xlsx"

df = pd.read_excel(excel_file, sheet_name="radb")

df["Engine_usage_max"] = pd.to_numeric(df["Engine_usage_max"], errors="coerce")
df = df.dropna(subset=["Engine_usage_max"])


def convert_to_km(row):
    motor = str(row["Motor"])
    engine_usage = row["Engine_usage_max"]

    if motor.startswith("5CT"):
        return engine_usage / 335000
    elif motor.startswith("OMC"):
        return engine_usage / 268000
    else:
        return None


def semaforo(km):
    if km >= SERVICE_LIMIT:
        return "🔴 SERVICE VENCIDO"
    elif km >= WARNING_LIMIT:
        return "🟡 PRÓXIMO SERVICE"
    else:
        return "🟢 OK"


df["km"] = df.apply(convert_to_km, axis=1)
df = df.dropna(subset=["km"])

kpi = df.groupby("Motor", as_index=False).agg(
    km_desde_ultimo_service=("km", "sum")
)

kpi["km_restantes"] = SERVICE_LIMIT - kpi["km_desde_ultimo_service"]
kpi["semaforo"] = kpi["km_desde_ultimo_service"].apply(semaforo)

services = (
    df[df["Engine_usage_max"] < 0]
    .groupby("Motor")
    .size()
    .reset_index(name="services_hechos")
)

kpi = kpi.merge(services, on="Motor", how="left")
kpi["services_hechos"] = kpi["services_hechos"].fillna(0).astype(int)

curr_state = pd.read_excel(excel_file, sheet_name="currState")
curr_state = curr_state[["Motor", "Piloto", "Equipo"]]

kpi = kpi.merge(curr_state, on="Motor", how="left")

kpi["km_desde_ultimo_service"] = kpi["km_desde_ultimo_service"].round().astype(int)
kpi["km_restantes"] = kpi["km_restantes"].round().astype(int)

def prioridad(estado):
    if "VENCIDO" in estado:
        return 1
    elif "PRÓXIMO" in estado:
        return 2
    else:
        return 3


kpi["prioridad"] = kpi["semaforo"].apply(prioridad)
kpi = kpi.sort_values(by=["prioridad", "km_restantes"], ascending=[True, True])
kpi = kpi.drop(columns=["prioridad"])

kpi = kpi.rename(
    columns={
        "km_desde_ultimo_service": "Km desde último Service",
        "km_restantes": "Km restantes",
        "services_hechos": "Services realizados",
        "semaforo": "Estado",
    }
)

kpi = kpi[
    [
        "Motor",
        "Piloto",
        "Equipo",
        "Km desde último Service",
        "Km restantes",
        "Services realizados",
        "Estado",
    ]
]

next_service = kpi[kpi["Km restantes"] > 0].sort_values("Km restantes").iloc[0]
max_usage = kpi.loc[kpi["Km desde último Service"].idxmax()]

summary = pd.DataFrame({
    "Indicador": [
        "Fecha de generación",
        "Motores monitoreados",
        "Service vencido",
        "Próximo service",
        "OK",
        "Motor con mayor uso",
        "Próximo motor a service",
    ],
    "Valor": [
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        len(kpi),
        (kpi["Estado"] == "🔴 SERVICE VENCIDO").sum(),
        (kpi["Estado"] == "🟡 PRÓXIMO SERVICE").sum(),
        (kpi["Estado"] == "🟢 OK").sum(),
        f'{max_usage["Motor"]} - {max_usage["Km desde último Service"]} km',
        f'{next_service["Motor"]} - {next_service["Km restantes"]} km restantes',
    ],
})

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    summary.to_excel(writer, sheet_name="Resumen Ejecutivo", index=False)
    kpi.to_excel(writer, sheet_name="Engine KPI Report", index=False)

wb = load_workbook(output_file)
ws = wb["Engine KPI Report"]

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for cell in ws[1]:
    cell.font = Font(bold=True)

estado_column = None

for cell in ws[1]:
    if cell.value == "Estado":
        estado_column = cell.column
        break

for row in range(2, ws.max_row + 1):
    estado_cell = ws.cell(row=row, column=estado_column)

    if "OK" in str(estado_cell.value):
        estado_cell.fill = green_fill
    elif "PRÓXIMO" in str(estado_cell.value):
        estado_cell.fill = yellow_fill
    elif "VENCIDO" in str(estado_cell.value):
        estado_cell.fill = red_fill

for column_cells in ws.columns:
    max_length = max(
        len(str(cell.value)) if cell.value is not None else 0
        for cell in column_cells
    )
    column_letter = get_column_letter(column_cells[0].column)
    ws.column_dimensions[column_letter].width = max_length + 2

wb.save(output_file)

print("✅ Motor Dashboard v1.1 generado correctamente.")
